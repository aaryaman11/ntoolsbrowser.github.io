import sys
import json
import openpyxl
from pathlib import Path
import itertools

def get_connection(index):
    connect_index = int(index) - 1
    connect = {
        "index": connect_index,
        "elecID": coor_dicts[connect_index]["elecID"],
    }
    return connect

def get_coords(x, y, z):
    coords = {
        "x": float(x),
        "y": float(y),
        "z": float(z),
    }
    return coords

# path to txt
coord_file = open(sys.argv[1], 'r')

# path to xlsx
attributes = Path(sys.argv[2])
fmaps = Path(sys.argv[3])
# fmaps = pd.ExcelFile(sys.argv[3])

subject_name = input("Enter Subject ID: ")

attr_wb = openpyxl.load_workbook(attributes)
attr_sheet = attr_wb.active
attr_rows = attr_sheet.iter_rows(values_only=True)
# remove title attrs
next(attr_rows)

fmap_wb = openpyxl.load_workbook(fmaps, enumerate)
fmap_sheet = fmap_wb.active
fmap_rows = fmap_sheet.iter_rows(values_only=True)
next(fmap_rows)

num_fmap_rows = fmap_sheet.max_row
fmap_dicts = []
num_cols = 10

content = coord_file.read()
coor_dicts = []
final_json = { "subjID" : subject_name, 
               "totalSeizType": 2,
               "SeizDisplay": ["seizType1", "seizType2", "intPopulation", "funMapping"]
 }

for line in content.split('\n'):
    data_tuple = tuple(line.rstrip().split(' '))
    try:
        (ID, x, y, z, electype) = data_tuple
        (ID2, intPop, seiz1, seiz2) = next(attr_rows)

        electrode_data = {
            "elecID": ID,
            "coordinates": get_coords(x, y, z),
            "elecType": electype,
            "intPopulation": intPop if intPop else 0,
            "seizType1": seiz1 or '',
            "seizType2": seiz2 or '',
        }

        coor_dicts.append(electrode_data)
        
    except ValueError:
        # happens when an extra newline is at the end of the file
        continue
        
final_json["electrodes"] = coor_dicts

for i in range(num_fmap_rows):
    fmap_data = {}
    try:  
        cleaned_data = next(fmap_rows)[:num_cols]
        if not cleaned_data[0]:
            break

        (cntct, G1, G2, thrshld, mtr, \
        snsry, lng, vsl, othr, dschrgs) = cleaned_data

        fmap_data = {
            "fmapContact": cntct,
            "fmapG1": get_connection(G1),
            "fmapG2": get_connection(G2),
            "fmapThreshold": int(thrshld) if thrshld else 0,
            "fmapMotor": mtr or '',
            "fmapSensory": snsry or '',
            "fmapLanguage": lng or '',
            "fmapVisual": vsl or '',
            "fmapOther": othr or '',
            "fmapAfterDischarge": dschrgs,
        }

        fmap_dicts.append(fmap_data)
    except ValueError:
        continue

final_json["functionalMaps"] = fmap_dicts
with open(f'sub-{subject_name}_ntoolsbrowser.json', 'w') as output_file:
    json.dump(final_json, output_file, ensure_ascii=False, indent=4)
